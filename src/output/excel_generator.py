"""
Liquidline Bank Reconciliation Automation
Excel Output Generator

Generates Excel files for:
1. Curtis's review - pre-populated cash posting data
2. Eagle import format - data ready for Eagle import
3. Erin's reconciliation view

Format matches Curtis's existing spreadsheet with added columns
for automation results.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Union
import logging

from ..models.transaction import Transaction, ConfidenceLevel, MatchResult

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generates Excel output files for the automation system

    Output includes:
    - Transaction details from bank download
    - Matched customer code and name
    - Confidence score and level (color-coded)
    - Invoice allocations (Karen's requirement)
    - Action recommendation (Auto/Review/Exception)
    """

    # Color coding for confidence levels
    COLORS = {
        "high": "#90EE90",      # Light green
        "medium": "#FFFF99",    # Light yellow
        "low": "#FFB6C1",       # Light pink/red
        "unmatched": "#D3D3D3"  # Light gray
    }

    def __init__(self, output_dir: Optional[Union[str, Path]] = None):
        """
        Args:
            output_dir: Directory to save output files
        """
        self.output_dir = Path(output_dir) if output_dir else Path("output")
        self.output_dir.mkdir(exist_ok=True)

    def generate_curtis_review(
        self,
        transactions: List[Transaction],
        filename: Optional[str] = None
    ) -> Path:
        """
        Generate Excel file for Curtis's review

        This is the main output - pre-populated data for Curtis to review
        and approve before posting to Eagle.

        Args:
            transactions: List of processed transactions
            filename: Output filename (auto-generated if not provided)

        Returns:
            Path to generated file
        """
        if not filename:
            date_str = datetime.now().strftime("%Y-%m-%d_%H%M")
            filename = f"Cash_Posting_Review_{date_str}.xlsx"

        output_path = self.output_dir / filename

        # Build data for export
        rows = []
        for t in transactions:
            row = self._build_curtis_row(t)
            rows.append(row)

        # Create DataFrame
        df = pd.DataFrame(rows)

        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Cash Posting', index=False)

            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Cash Posting']

            # Apply formatting
            self._apply_formatting(worksheet, df)

            # Add summary sheet
            self._add_summary_sheet(writer, transactions)

        logger.info(f"Generated Curtis review file: {output_path}")
        return output_path

    def _build_curtis_row(self, transaction: Transaction) -> dict:
        """Build a row for Curtis's review spreadsheet"""
        match = transaction.match_result

        # Determine action recommendation
        if match:
            if match.confidence_level == ConfidenceLevel.HIGH:
                action = "AUTO"
            elif match.confidence_level == ConfidenceLevel.MEDIUM:
                action = "REVIEW"
            else:
                action = "EXCEPTION"
        else:
            action = "MANUAL"

        # Build invoice allocation string
        allocations_str = ""
        if match and match.invoice_allocations:
            alloc_parts = []
            for alloc in match.invoice_allocations:
                alloc_parts.append(f"{alloc.invoice_number}: £{alloc.allocated_amount:,.2f}")
            allocations_str = "; ".join(alloc_parts)

        return {
            # Bank data
            "Row": transaction.row_id,
            "Post Date": transaction.post_date.strftime("%d/%m/%Y"),
            "Type": transaction.transaction_type,
            "Amount": transaction.amount,
            "Customer Reference": transaction.customer_reference,
            "Transaction Detail": transaction.transaction_detail,

            # Matching results
            "Matched Customer Code": match.customer_code if match else "",
            "Matched Customer Name": match.customer_name if match else "",
            "Confidence": f"{match.confidence_score * 100:.0f}%" if match else "0%",
            "Confidence Level": match.confidence_level.value.upper() if match else "UNMATCHED",
            "Match Method": match.match_method.value if match else "",

            # Invoice allocations (Karen's requirement)
            "Invoice Allocations": allocations_str,

            # Action
            "Recommended Action": action,
            "Match Details": match.match_details if match else "No match found",

            # Manual fields (for Curtis to complete)
            "Posted?": "",
            "Posted To": "",
            "Comments": "",

            # Source tracking
            "Source File": transaction.source_file
        }

    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply conditional formatting to worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define fills for confidence levels
        fills = {
            "HIGH": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
            "LOW": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
            "UNMATCHED": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        }

        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Format data rows
        confidence_col = df.columns.get_loc("Confidence Level") + 1

        for row_num in range(2, len(df) + 2):
            # Get confidence level for this row
            conf_cell = worksheet.cell(row=row_num, column=confidence_col)
            conf_value = conf_cell.value

            # Apply fill based on confidence
            if conf_value in fills:
                fill = fills[conf_value]
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    cell.border = thin_border

        # Set column widths
        column_widths = {
            "Row": 6,
            "Post Date": 12,
            "Type": 18,
            "Amount": 12,
            "Customer Reference": 25,
            "Transaction Detail": 30,
            "Matched Customer Code": 20,
            "Matched Customer Name": 25,
            "Confidence": 12,
            "Confidence Level": 15,
            "Match Method": 15,
            "Invoice Allocations": 35,
            "Recommended Action": 18,
            "Match Details": 40,
            "Posted?": 8,
            "Posted To": 15,
            "Comments": 25,
            "Source File": 25
        }

        for col_num, column_title in enumerate(df.columns, 1):
            width = column_widths.get(column_title, 15)
            worksheet.column_dimensions[get_column_letter(col_num)].width = width

        # Freeze top row
        worksheet.freeze_panes = "A2"

    def _add_summary_sheet(self, writer, transactions: List[Transaction]):
        """Add summary statistics sheet"""
        # Calculate stats
        total = len(transactions)
        matched = sum(1 for t in transactions if t.match_result)
        high_conf = sum(1 for t in transactions if t.match_result and t.match_result.confidence_level == ConfidenceLevel.HIGH)
        medium_conf = sum(1 for t in transactions if t.match_result and t.match_result.confidence_level == ConfidenceLevel.MEDIUM)
        low_conf = sum(1 for t in transactions if t.match_result and t.match_result.confidence_level == ConfidenceLevel.LOW)
        unmatched = total - matched

        total_amount = sum(t.amount for t in transactions)
        high_amount = sum(t.amount for t in transactions if t.match_result and t.match_result.confidence_level == ConfidenceLevel.HIGH)

        summary_data = [
            ["Liquidline Cash Posting Automation - Summary", ""],
            ["Generated", datetime.now().strftime("%d/%m/%Y %H:%M")],
            ["", ""],
            ["Total Transactions", total],
            ["Total Amount", f"£{total_amount:,.2f}"],
            ["", ""],
            ["Matching Results", ""],
            ["High Confidence (GREEN)", f"{high_conf} ({high_conf/total*100:.1f}%)"],
            ["Medium Confidence (YELLOW)", f"{medium_conf} ({medium_conf/total*100:.1f}%)"],
            ["Low Confidence (RED)", f"{low_conf} ({low_conf/total*100:.1f}%)"],
            ["Unmatched (GRAY)", f"{unmatched} ({unmatched/total*100:.1f}%)"],
            ["", ""],
            ["Automation Rate", f"{matched/total*100:.1f}%"],
            ["High Confidence Amount", f"£{high_amount:,.2f} ({high_amount/total_amount*100:.1f}%)"],
            ["", ""],
            ["Recommended Actions", ""],
            ["AUTO (ready to post)", high_conf],
            ["REVIEW (check match)", medium_conf],
            ["EXCEPTION (manual)", low_conf + unmatched],
        ]

        summary_df = pd.DataFrame(summary_data, columns=["Metric", "Value"])
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)

    def generate_eagle_import(
        self,
        transactions: List[Transaction],
        filename: Optional[str] = None
    ) -> Path:
        """
        Generate Eagle import format file

        This format is specifically for importing into Eagle ERP.
        CRITICAL: Only includes matches WITH customer_code (postable to Eagle).
        Matches without customer_code go to "Needs Review" sheet.

        Args:
            transactions: List of processed transactions
            filename: Output filename

        Returns:
            Path to generated file
        """
        if not filename:
            date_str = datetime.now().strftime("%Y-%m-%d_%H%M")
            filename = f"Eagle_Import_{date_str}.xlsx"

        output_path = self.output_dir / filename

        # Split transactions by postability (has customer_code)
        ready_to_post = []
        needs_review = []

        for t in transactions:
            if not t.match_result:
                continue
            match = t.match_result
            # CRITICAL: Only include if we have a customer code
            has_customer_code = bool(match.customer_code and match.customer_code.strip())

            row_data = self._build_eagle_row(t, match)

            if has_customer_code and match.confidence_level == ConfidenceLevel.HIGH:
                ready_to_post.append(row_data)
            else:
                needs_review.append(row_data)

        ready_df = pd.DataFrame(ready_to_post) if ready_to_post else pd.DataFrame()
        review_df = pd.DataFrame(needs_review) if needs_review else pd.DataFrame()

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Ready to Post sheet - these can go directly into Eagle
            if not ready_df.empty:
                ready_df.to_excel(writer, sheet_name='Ready to Post', index=False)
            else:
                pd.DataFrame({"Message": ["No transactions ready to post - all need review"]}).to_excel(
                    writer, sheet_name='Ready to Post', index=False)

            # Needs Review sheet - missing customer code or low confidence
            if not review_df.empty:
                review_df.to_excel(writer, sheet_name='Needs Review', index=False)

            # Add summary
            summary_data = [
                ["Eagle Import Summary", ""],
                ["Generated", datetime.now().strftime("%d/%m/%Y %H:%M")],
                ["", ""],
                ["Ready to Post (has customer code)", len(ready_to_post)],
                ["Needs Review (missing customer code)", len(needs_review)],
                ["", ""],
                ["Total Matched", len(ready_to_post) + len(needs_review)],
                ["Postable Rate", f"{len(ready_to_post)/(len(ready_to_post)+len(needs_review))*100:.1f}%" if (len(ready_to_post)+len(needs_review)) > 0 else "0%"],
            ]
            pd.DataFrame(summary_data, columns=["Metric", "Value"]).to_excel(
                writer, sheet_name='Summary', index=False, header=False)

        logger.info(f"Generated Eagle import file: {output_path} ({len(ready_to_post)} postable, {len(needs_review)} need review)")
        return output_path

    def _build_eagle_row(self, t: Transaction, match: MatchResult) -> dict:
        """Build a single row for Eagle export"""
        # Create row for primary invoice or transaction
        if match.invoice_allocations and len(match.invoice_allocations) > 0:
            alloc = match.invoice_allocations[0]
            return {
                "Post Date": t.post_date.strftime("%d/%m/%Y"),
                "Customer Code": match.customer_code,
                "Customer Name": match.customer_name,
                "Amount": t.amount,
                "Invoice Number": alloc.invoice_number,
                "All Invoices": "; ".join([a.invoice_number for a in match.invoice_allocations]),
                "Bank Reference": t.customer_reference,
                "Payment Method": self._map_payment_method(t.transaction_type),
                "Confidence": f"{match.confidence_score * 100:.0f}%",
                "Needs Lookup": "YES" if not match.customer_code else "",
                "Notes": match.match_details
            }
        else:
            return {
                "Post Date": t.post_date.strftime("%d/%m/%Y"),
                "Customer Code": match.customer_code,
                "Customer Name": match.customer_name,
                "Amount": t.amount,
                "Invoice Number": "",
                "All Invoices": "",
                "Bank Reference": t.customer_reference,
                "Payment Method": self._map_payment_method(t.transaction_type),
                "Confidence": f"{match.confidence_score * 100:.0f}%",
                "Needs Lookup": "YES" if not match.customer_code else "",
                "Notes": match.match_details
            }

    def _map_payment_method(self, transaction_type: str) -> str:
        """Map bank transaction type to Eagle payment method"""
        mapping = {
            "Bank Giro Credit": "BACS",
            "Faster Payment": "FP",
            "CHAPS": "CHAPS",
            "Direct Debit": "DD",
            "BACS": "BACS"
        }
        return mapping.get(transaction_type, "OTHER")
